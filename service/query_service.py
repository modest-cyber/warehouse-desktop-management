# -*- coding: utf-8 -*-
"""
查询统计Service
实现各种查询和统计功能
"""

from typing import List, Optional, Dict, Any
from datetime import datetime
from dao.stock_dao import StockDAO
from dao.inventory_dao import InventoryDAO
from dao.product_dao import ProductDAO
from dao.warehouse_dao import WarehouseDAO
from dao.supplier_client_dao import SupplierClientDAO
from model.stock_record import StockRecord
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from utils.logger import Logger

logger = Logger.get_logger(__name__)


class QueryService:
    """查询统计Service类"""
    
    def __init__(self):
        """初始化Service"""
        self.stock_dao = StockDAO()
        self.inventory_dao = InventoryDAO()
        self.product_dao = ProductDAO()
        self.warehouse_dao = WarehouseDAO()
        self.supplier_client_dao = SupplierClientDAO()
    
    def query_stock_records(self, conditions: Dict[str, Any]) -> List[StockRecord]:
        """
        查询出入库记录（支持多条件组合）
        
        Args:
            conditions: 查询条件字典，可包含：
                - warehouse_id: 仓库ID
                - product_id: 货品ID
                - supplier_client_id: 供应商/客户ID
                - record_type: 记录类型（1-入库，2-出库）
                - start_date: 开始日期
                - end_date: 结束日期
                - record_no: 单据号（模糊查询）
                - operator: 操作人（模糊查询）
            
        Returns:
            list: 出入库记录对象列表
        """
        sql = """
            SELECT sr.* FROM stock_record sr WHERE 1=1
        """
        params = []
        
        if conditions.get('warehouse_id'):
            sql += " AND sr.warehouse_id=%s"
            params.append(conditions['warehouse_id'])
        
        if conditions.get('product_id'):
            sql += " AND sr.product_id=%s"
            params.append(conditions['product_id'])
        
        if conditions.get('supplier_client_id'):
            sql += " AND sr.supplier_client_id=%s"
            params.append(conditions['supplier_client_id'])
        
        if conditions.get('record_type'):
            sql += " AND sr.record_type=%s"
            params.append(conditions['record_type'])
        
        if conditions.get('start_date'):
            sql += " AND sr.record_date >= %s"
            params.append(conditions['start_date'])
        
        if conditions.get('end_date'):
            sql += " AND sr.record_date <= %s"
            params.append(conditions['end_date'])
        
        if conditions.get('record_no'):
            sql += " AND sr.record_no LIKE %s"
            params.append(f"%{conditions['record_no']}%")
        
        if conditions.get('operator'):
            sql += " AND sr.operator LIKE %s"
            params.append(f"%{conditions['operator']}%")
        
        sql += " ORDER BY sr.record_date DESC, sr.id DESC"
        
        results = self.stock_dao.fetch_all(sql, tuple(params) if params else None)
        return [StockRecord.from_dict(row) for row in results]
    
    def query_stock_records_by_date(self, start_date: datetime, end_date: datetime) -> List[StockRecord]:
        """
        按日期范围查询
        
        Args:
            start_date: 开始日期
            end_date: 结束日期
            
        Returns:
            list: 出入库记录对象列表
        """
        return self.query_stock_records({'start_date': start_date, 'end_date': end_date})
    
    def query_stock_records_by_warehouse(self, warehouse_id: int, 
                                        start_date: Optional[datetime] = None,
                                        end_date: Optional[datetime] = None) -> List[StockRecord]:
        """
        按仓库查询
        
        Args:
            warehouse_id: 仓库ID
            start_date: 开始日期（可选）
            end_date: 结束日期（可选）
            
        Returns:
            list: 出入库记录对象列表
        """
        conditions = {'warehouse_id': warehouse_id}
        if start_date:
            conditions['start_date'] = start_date
        if end_date:
            conditions['end_date'] = end_date
        return self.query_stock_records(conditions)
    
    def query_stock_records_by_product(self, product_id: int,
                                      start_date: Optional[datetime] = None,
                                      end_date: Optional[datetime] = None) -> List[StockRecord]:
        """
        按货品查询
        
        Args:
            product_id: 货品ID
            start_date: 开始日期（可选）
            end_date: 结束日期（可选）
            
        Returns:
            list: 出入库记录对象列表
        """
        conditions = {'product_id': product_id}
        if start_date:
            conditions['start_date'] = start_date
        if end_date:
            conditions['end_date'] = end_date
        return self.query_stock_records(conditions)
    
    def query_stock_records_by_supplier_client(self, supplier_client_id: int,
                                              start_date: Optional[datetime] = None,
                                              end_date: Optional[datetime] = None) -> List[StockRecord]:
        """
        按供应商/客户查询
        
        Args:
            supplier_client_id: 供应商/客户ID
            start_date: 开始日期（可选）
            end_date: 结束日期（可选）
            
        Returns:
            list: 出入库记录对象列表
        """
        conditions = {'supplier_client_id': supplier_client_id}
        if start_date:
            conditions['start_date'] = start_date
        if end_date:
            conditions['end_date'] = end_date
        return self.query_stock_records(conditions)
    
    def get_inventory_statistics(self, warehouse_id: Optional[int] = None) -> Dict[str, Any]:
        """
        获取库存统计
        
        Args:
            warehouse_id: 仓库ID（可选，None表示所有仓库）
            
        Returns:
            dict: 统计结果，包含：
                - total_quantity: 总库存数量
                - total_value: 总库存价值
                - product_count: 货品种类数
                - low_stock_count: 低库存货品数
                - over_stock_count: 超库存货品数
        """
        if warehouse_id:
            inventories = self.inventory_dao.get_by_warehouse(warehouse_id)
        else:
            inventories = self.inventory_dao.get_all()
        
        total_quantity = 0
        total_value = 0
        product_count = len(set(inv.product_id for inv in inventories))
        low_stock_count = 0
        over_stock_count = 0
        
        for inventory in inventories:
            total_quantity += inventory.quantity
            product = self.product_dao.get_by_id(inventory.product_id)
            if product and product.price:
                total_value += float(inventory.quantity * product.price)
            
            if product:
                if inventory.quantity < product.min_stock:
                    low_stock_count += 1
                if product.max_stock > 0 and inventory.quantity > product.max_stock:
                    over_stock_count += 1
        
        return {
            'total_quantity': total_quantity,
            'total_value': round(total_value, 2),
            'product_count': product_count,
            'low_stock_count': low_stock_count,
            'over_stock_count': over_stock_count
        }
    
    def get_product_statistics(self, product_id: int) -> Dict[str, Any]:
        """
        获取货品统计
        
        Args:
            product_id: 货品ID
            
        Returns:
            dict: 统计结果
        """
        inventories = self.inventory_dao.get_by_product(product_id)
        product = self.product_dao.get_by_id(product_id)
        
        total_quantity = sum(inv.quantity for inv in inventories)
        total_value = 0
        if product and product.price:
            total_value = float(total_quantity * product.price)
        
        return {
            'product_id': product_id,
            'product_name': product.product_name if product else '',
            'total_quantity': total_quantity,
            'total_value': round(total_value, 2),
            'warehouse_count': len(inventories)
        }
    
    def get_warehouse_statistics(self, warehouse_id: int) -> Dict[str, Any]:
        """
        获取仓库统计
        
        Args:
            warehouse_id: 仓库ID
            
        Returns:
            dict: 统计结果
        """
        inventories = self.inventory_dao.get_by_warehouse(warehouse_id)
        warehouse = self.warehouse_dao.get_by_id(warehouse_id)
        
        total_quantity = sum(inv.quantity for inv in inventories)
        total_value = 0
        for inventory in inventories:
            product = self.product_dao.get_by_id(inventory.product_id)
            if product and product.price:
                total_value += float(inventory.quantity * product.price)
        
        return {
            'warehouse_id': warehouse_id,
            'warehouse_name': warehouse.warehouse_name if warehouse else '',
            'total_quantity': total_quantity,
            'total_value': round(total_value, 2),
            'product_count': len(inventories)
        }
    
    def get_supplier_client_statistics(self, supplier_client_id: int,
                                      start_date: Optional[datetime] = None,
                                      end_date: Optional[datetime] = None) -> Dict[str, Any]:
        """
        获取供应商/客户往来统计
        
        Args:
            supplier_client_id: 供应商/客户ID
            start_date: 开始日期（可选）
            end_date: 结束日期（可选）
            
        Returns:
            dict: 统计结果
        """
        conditions = {'supplier_client_id': supplier_client_id}
        if start_date:
            conditions['start_date'] = start_date
        if end_date:
            conditions['end_date'] = end_date
        
        records = self.query_stock_records(conditions)
        supplier_client = self.supplier_client_dao.get_by_id(supplier_client_id)
        
        in_amount = 0
        out_amount = 0
        in_quantity = 0
        out_quantity = 0
        
        for record in records:
            if record.record_type == 1:  # 入库
                in_quantity += record.quantity
                if record.total_amount:
                    in_amount += float(record.total_amount)
            else:  # 出库
                out_quantity += record.quantity
                if record.total_amount:
                    out_amount += float(record.total_amount)
        
        return {
            'supplier_client_id': supplier_client_id,
            'name': supplier_client.name if supplier_client else '',
            'in_quantity': in_quantity,
            'out_quantity': out_quantity,
            'in_amount': round(in_amount, 2),
            'out_amount': round(out_amount, 2),
            'net_amount': round(in_amount - out_amount, 2)
        }
    
    def export_to_excel(self, data: List[Dict[str, Any]], file_path: str, 
                       sheet_name: str = 'Sheet1', headers: Optional[List[str]] = None) -> bool:
        """
        导出数据到Excel
        
        Args:
            data: 要导出的数据列表（字典列表）
            file_path: 文件路径
            sheet_name: Sheet名称
            headers: 表头列表（如果为None，则使用data中第一个字典的键）
            
        Returns:
            bool: 是否导出成功
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # 设置表头
            if headers:
                header_row = headers
            elif data:
                header_row = list(data[0].keys())
            else:
                return False
            
            # 写入表头
            for col, header in enumerate(header_row, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # 写入数据
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(header_row, 1):
                    value = row_data.get(header, '')
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # 保存文件
            wb.save(file_path)
            logger.info(f"导出Excel成功: {file_path}")
            return True
        except Exception as e:
            logger.error(f"导出Excel失败: {str(e)}")
            raise Exception(f"导出Excel失败: {str(e)}")

